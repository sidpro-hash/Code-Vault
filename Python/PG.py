import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font
from openpyxl.styles.borders import Border, Side

year = int(2022)
month = int(2)

'''change day and respective columns to map'''
day = int(7)
emi_column1 = 'P%d'
emi_column2 = 'Q%d'
track_column = 'J%d'

'''change value of filename or sheet_name'''
df = pd.read_excel(r'C:/Users/GabuSiddharthMerambh/Downloads/CONSOLIDATED PRESENATION MASTER DATA.xlsx', engine='openpyxl',sheet_name='FEB.22')
df.head()
# Presentation Mode
#  Final Remark
#  Clearence received Date
#  EMI COLLECTED MODE
# EMI received date
paid = df[df["Final Remark"]=='PAID']
no_mode_with_razorpay = paid[(paid["Presentation Mode"]=='CASH') | (paid["Presentation Mode"]=='NEFT') | (paid["Presentation Mode"]=='RAZORPAY')]
no_mode_without_razorpay = paid[(paid["Presentation Mode"]=='CASH') | (paid["Presentation Mode"]=='NEFT')]
no_mode_A = len(no_mode_with_razorpay.index)
no_mode_B = len(no_mode_without_razorpay.index)
# year,month,day
# pd.Timestamp(2022,1,13)
# year,month,day
Clearence_received = no_mode_with_razorpay[no_mode_with_razorpay["Clearence received Date"]==pd.Timestamp(year,month,day)]["Loan Account number"]
Clearence_received.to_csv("Clearence_received.csv", index=False)
mode_with_razorpay = paid[((paid["Presentation Mode"]=='ESCROW') | (paid["Presentation Mode"]=='NACH') | (paid["Presentation Mode"]=='PDC_BRANCH') | (paid["Presentation Mode"]=='PDC_CENTRAL OPS')) & (paid["EMI COLLECTED MODE"]!='-') & (paid["EMI COLLECTED MODE"].str[-7:]!='Cleared')]
mode_without_razorpay = mode_with_razorpay[(mode_with_razorpay["EMI COLLECTED MODE"].str[-18:]!='RAZORPAY COLLECTED') & (mode_with_razorpay["EMI COLLECTED MODE"]!='RAZORPAY')]
mode_A = len(mode_with_razorpay.index)
mode_B = len(mode_without_razorpay.index)
# year,month,day
EMI_received = mode_with_razorpay[mode_with_razorpay["EMI received date"]==pd.Timestamp(year,month,day)]["Loan Account number"]
final = pd.DataFrame({"No Mode A":no_mode_A,"No Mode B":no_mode_B,"Mode A":mode_A,"Mode B":mode_B,"No Mode Clearence received":Clearence_received,"Mode EMI received":EMI_received})
final.to_excel("foo.xlsx", sheet_name="Sheet1")


Clearence_received = pd.DataFrame(Clearence_received)
EMI_received = pd.DataFrame(EMI_received)

#styles
fill_cell1 = PatternFill(patternType='solid',fgColor='9bfc92')
font1 = Font(color='19960e')
fill_cell2 = PatternFill(patternType='solid',fgColor='cccccc')
fill_cell3 = PatternFill(patternType='solid',fgColor='ffff00')
font2 = Font(color='595959',bold=True)
font3 = Font(color='914e00')
fill_cell4 = PatternFill(patternType='solid',fgColor='fdff78')
thin_border = Border(left=Side(border_style='thin',color='544a49'),right=Side(border_style='thin',color='544a49'),top=Side(border_style='thin',color='544a49'),bottom=Side(border_style='thin',color='544a49'))
_border = Border(left=Side(border_style='thin'),right=Side(border_style='thin'),top=Side(border_style='thin'),bottom=Side(border_style='thin'))


wb = load_workbook('foo.xlsx')
ws = wb['Sheet1']
emi_list = []
cler_list = []
for i in range(4):
    cell = emi_column1 % (i+3) 
    cell1 = emi_column2 % (i+3)
    emi_list.append(cell)
    cler_list.append(cell1)
    #print(emi_list[i])


ws[emi_list[0]] = pd.Timestamp(year,month,day).to_pydatetime()
ws[emi_list[0]].value = ws[emi_list[0]].value.strftime('%Y-%m-%d')
ws[emi_list[0]].fill = fill_cell4
ws[emi_list[0]].font = font3
ws[emi_list[1]] = 'EMI received'
ws[emi_list[1]].fill = fill_cell1
ws[emi_list[1]].font = font1
ws[emi_list[1]].border = thin_border
ws[emi_list[2]] = 'MODE'
ws[emi_list[2]].fill = fill_cell2
ws[emi_list[2]].font = font2
ws[emi_list[2]].border = thin_border
ws[emi_list[3]] = 'Loan Account number'
ws[emi_list[3]].fill = fill_cell3
ws[emi_list[3]].border = _border


ws[cler_list[1]] = 'Clearence received'
ws[cler_list[1]].fill = fill_cell1
ws[cler_list[1]].font = font1
ws[cler_list[1]].border = thin_border
ws[cler_list[2]] = 'NO MODE'
ws[cler_list[2]].fill = fill_cell2
ws[cler_list[2]].font = font2
ws[cler_list[2]].border = thin_border
ws[cler_list[3]] = 'Loan Account number'
ws[cler_list[3]].fill = fill_cell3
ws[cler_list[3]].border = _border



i=0
for index,row in EMI_received.iterrows():
    cell = emi_column1 % (i+7)
    #print(cell)
    i+=1
    ws[cell] = row[0]
    ws[cell].border = _border
    #print(ws[cell])

i=0
for index,row in Clearence_received.iterrows():
    cell = emi_column2 % (i+7)
    #print(cell)
    i+=1
    ws[cell] = row[0]
    ws[cell].border = thin_border
    #print(ws[cell])
    
wb.save('foo.xlsx')

track_list = []
for i in range(4):
    cell = track_column % (i+6)
    track_list.append(cell)

wb = load_workbook('test.xlsx')
ws = wb['Sheet1'] 
ws[track_list[0]]=mode_A
ws[track_list[1]]=mode_B
ws[track_list[2]]=no_mode_A
ws[track_list[3]]=no_mode_B

wb.save('test.xlsx')